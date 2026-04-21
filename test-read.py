import openpyxl

wb = openpyxl.load_workbook('test_output.xlsx')
ws = wb.active
print(f"C10: {ws['C10'].value}, type: {type(ws['C10'].value)}")
print(f"A12: {ws['A12'].value}, type: {type(ws['A12'].value)}")
